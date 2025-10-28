import streamlit as st
import pandas as pd
import re
from openpyxl import load_workbook

# 页面配置
st.set_page_config(
    page_title="公摊对数专用",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# 标题与说明
st.title("公摊对数专用")
st.write("兼容所有Excel版本，自动获取表格2公式计算结果")
st.divider()

# 初始化会话状态
if 'table1' not in st.session_state:
    st.session_state.table1 = None
if 'table2' not in st.session_state:
    st.session_state.table2 = {}
if 'result' not in st.session_state:
    st.session_state.result = pd.DataFrame(columns=[
        "资源名称", "增量推账金额(元)", "表单-楼层", "总计每户分摊金额", "对比结果"
    ])

# 1. 文件上传区
st.subheader("1. 上传表格文件")
col1, col2 = st.columns(2)

with col1:
    file1 = st.file_uploader("上传表格1（含账单数据）", type=["xlsx", "xls"])
    if file1:
        try:
            df = pd.read_excel(file1, sheet_name="账单数据")
            if "资源名称" in df.columns and "增量推账金额(元)" in df.columns:
                df["增量推账金额(元)"] = pd.to_numeric(df["增量推账金额(元)"], errors="coerce")
                st.session_state.table1 = df
                st.success(f"表格1加载成功（{len(df)}行数据）")
                
                # 自动执行表格1数据提取
                df_extracted = df[["资源名称", "增量推账金额(元)"]].copy()
                df_extracted = df_extracted.dropna(subset=["资源名称"])
                df_extracted["增量推账金额(元)"] = df_extracted["增量推账金额(元)"].round(2)
                df_extracted["表单-楼层"] = ""
                df_extracted["总计每户分摊金额"] = pd.Series(dtype='float64')
                df_extracted["对比结果"] = ""
                st.session_state.result = df_extracted
                st.success(f"已自动提取表格1数据（{len(df_extracted)}行）")
            else:
                st.error("表格1缺少必要列：资源名称 / 增量推账金额(元)")
        except Exception as e:
            st.error(f"表格1读取失败：{str(e)}")

with col2:
    # 表格2上传按钮默认禁用，表格1上传成功后启用
    file2 = st.file_uploader(
        "上传表格2（含公式）", 
        type=["xlsx"],
        disabled=st.session_state.table1 is None
    )
    if file2:
        try:
            wb = load_workbook(file2, data_only=True)
            sheets = {}
            for sheet_name in wb.sheetnames:
                if sheet_name.endswith("栋") and sheet_name[:-1].isdigit():
                    bld_num = int(sheet_name[:-1])
                    if 1 <= bld_num <= 34:
                        ws = wb[sheet_name]
                        data = []
                        # 从第12行开始读取有效数据（前11行为无效表头）
                        for row in ws.iter_rows(min_row=12, values_only=True):
                            if len(row) >= 7:
                                floor = row[1]
                                amount = row[6]
                                if floor and amount is not None:
                                    data.append({"楼层": floor, "总计每户分摊金额": amount})
                        if data:
                            df = pd.DataFrame(data)
                            df["总计每户分摊金额"] = pd.to_numeric(df["总计每户分摊金额"], errors="coerce")
                            sheets[bld_num] = df
            st.session_state.table2 = sheets
            st.success(f"表格2加载成功（已读取公式结果，{len(sheets)}个楼栋表单）")
            
            # 上传表格2后自动执行匹配操作
            if not st.session_state.result.empty and st.session_state.table2:
                df = st.session_state.result.copy()
                for i, row in df.iterrows():
                    name = str(row["资源名称"])
                    bld_match = re.search(r"住宅(\d{2})-", name)
                    flr_match = re.search(r"-(\d{4})", name)
                    if bld_match and flr_match:
                        bld = int(bld_match.group(1))
                        flr_code = flr_match.group(1)
                        flr = f"{int(flr_code[:2])}层"  # 无空格格式
                        df.at[i, "表单-楼层"] = f"表单'{bld}栋'-{flr}"
                        if bld in st.session_state.table2:
                            bld_df = st.session_state.table2[bld]
                            mask = bld_df["楼层"].astype(str) == flr
                            if mask.any():
                                amount = bld_df[mask].iloc[0]["总计每户分摊金额"]
                                df.at[i, "总计每户分摊金额"] = round(float(amount), 2) if pd.notna(amount) else None
                st.session_state.result = df
                st.success("已自动完成表格2数据匹配（保留两位小数）")
            else:
                st.error("表格1数据为空，无法进行匹配")
                
        except Exception as e:
            st.error(f"表格2处理失败：{str(e)}")

st.divider()

# 2. 功能按钮区
st.subheader("2. 数据处理")
if not st.session_state.result.empty and st.session_state.table2:
    if st.button("对比数据", width='stretch'):
        df = st.session_state.result.copy()
        same_count = 0
        diff_count = 0
        total_count = len(df)
        for i, row in df.iterrows():
            try:
                amt1 = round(float(row["增量推账金额(元)"]), 2) if pd.notna(row["增量推账金额(元)"]) else None
                amt2 = round(float(row["总计每户分摊金额"]), 2) if pd.notna(row["总计每户分摊金额"]) else None
                if amt1 is not None and amt2 is not None:
                    if amt1 == amt2:
                        df.at[i, "对比结果"] = "一致"
                        same_count += 1
                    else:
                        df.at[i, "对比结果"] = "不一致"
                        diff_count += 1
                else:
                    df.at[i, "对比结果"] = "数据缺失"
            except:
                df.at[i, "对比结果"] = "格式错误"
        st.session_state.result = df
        st.success(f"对比完成！共 {total_count} 行数据")
        col_summary1, col_summary2, col_summary3 = st.columns(3)
        with col_summary1:
            st.info(f"一致：{same_count} 行（{round(same_count/total_count*100, 1)}%）")
        with col_summary2:
            st.warning(f"不一致：{diff_count} 行（{round(diff_count/total_count*100, 1)}%）")
        with col_summary3:
            other_count = total_count - same_count - diff_count
            st.error(f"其他（缺失/错误）：{other_count} 行")
else:
    st.info("请先完成表格1提取和表格2上传")

st.divider()

# 3. 结果展示区
st.subheader("3. 处理结果")
if not st.session_state.result.empty:
    def highlight(row):
        # 基础样式：第二列和第四列居中并设置高亮
        styles = [
            "", 
            "text-align: center; background-color: #FFFFFF;", 
            "", 
            "text-align: center; background-color: #FFFFFF;", 
            ""
        ]
        # 叠加对比结果的高亮样式
        if row["对比结果"] == "一致":
            styles[1] = "text-align: center; background-color: #90EE90;"
            styles[3] = "text-align: center; background-color: #90EE90;"
            styles[4] = "background-color: #90EE90;"
        elif row["对比结果"] == "不一致":
            styles[1] = "text-align: center; background-color: #FFA07A;"
            styles[3] = "text-align: center; background-color: #FFA07A;"
            styles[4] = "background-color: #FFA07A;"
        return styles
    
    def format_numbers(val):
        if pd.api.types.is_numeric_dtype(type(val)):
            return f"{val:.2f}" if not pd.isna(val) else ""
        return val
    
    styled_df = st.session_state.result.style.apply(highlight, axis=1)\
        .format({"增量推账金额(元)": format_numbers, "总计每户分摊金额": format_numbers})
    st.dataframe(styled_df, width='stretch', height=500)
else:
    st.info("请按步骤处理数据以显示结果")
